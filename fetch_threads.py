import requests
import json
import time
import re
from datetime import datetime, timezone, timedelta
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("openpyxl is not installed. Run:  pip3 install openpyxl")
    raise SystemExit(1)

# ─────────────────────────────────────────────────────────
#  CONFIGURATION
# ─────────────────────────────────────────────────────────
API_KEY = "YOUR_API_KEY_HERE"

# ── BATCH OF INSURERS ────────────────────────────────────
#   name    → full name shown in the scoring sheet + sheet tab name
#   handles → ALL handles: tagged by customers OR used to reply
#             (brand handle + care/support/group handles)
INSURERS = [
    {"name": "Aditya Birla Capital", "handles": ["abcapital", "abhealthin"]},
    {"name": "Aviva",                "handles": ["avivaindia"]},
    # ... add the rest of the batch
]

# ── DATE WINDOW ──────────────────────────────────────────
# since is INCLUSIVE, until is EXCLUSIVE. Midnight IST.
SEARCH_SINCE = "2026-04-01"
SEARCH_UNTIL = "2026-07-01"

# ── PRIORITY KEYWORDS (soft ranking) ─────────────────────
PRIORITY_KEYWORDS = [
    "term insurance", "term plan", "term policy",
    "claim", "settlement",
    "mis-sold", "missold", "mis-selling", "misselling", "fraud",
    "cancel", "surrender", "free look", "freelook",
    "lapse", "revival", "renewal", "refund",
    "premium", "nominee", "policy document",
]

ROOT_TWEETS_ONLY = True

# ── MEDIA / SPONSOR NOISE FILTER ─────────────────────────
MAX_AUTHOR_FOLLOWERS = 100000     # skip authors above this; None disables
SKIP_BUSINESS_VERIFIED = True     # skip gold/business verified authors
EXCLUDE_AUTHORS = ["revsportzglobal", "boriamajumdar", "cnbc_awaaz",
                   "cnbctv18news", "cnbctv18live", "etnowlive",
                   "etnowswadesh", "tv9kannada", "tv9telugu",
                   "tv9gujarati", "tv9marathi", "tv9bharatvarsh",
                   "thequint", "businessline", "tellychakkar"]

# ── LIMITS ───────────────────────────────────────────────
MAX_SEARCH_TWEETS  = 300   # per-insurer cap on complaints processed
MAX_SEARCH_PAGES   = 30    # advanced-search pages per insurer
MAX_TIMELINE_PAGES = 50    # insurer own-timeline pages per handle
MAX_THREAD_PAGES   = 5     # thread-context pages per thread

# ── OUTPUT FILES (one sheet per insurer) ─────────────────
OUTPUT_URLS    = "twitter_found_urls.xlsx"
OUTPUT_MAIN    = "twitter_scoring.xlsx"
OUTPUT_DETAIL  = "twitter_timestamps.xlsx"
OUTPUT_JSON    = "twitter_threads.json"

MAX_TIMESTAMPS = 9
# ─────────────────────────────────────────────────────────

IST  = timezone(timedelta(hours=5, minutes=30))
WAIT = 7

HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
BOLD = Font(bold=True)

# ──────────────── time helpers ───────────────────────────

def date_to_unix(date_str):
    dt = datetime.strptime(date_str, "%Y-%m-%d").replace(tzinfo=IST)
    return int(dt.timestamp())

def to_ist(ts_str):
    if not ts_str: return "N/A"
    try:
        dt = datetime.strptime(ts_str, "%a %b %d %H:%M:%S +0000 %Y")
        return dt.replace(tzinfo=timezone.utc).astimezone(IST).strftime("%d-%m-%Y %H:%M:%S")
    except: pass
    try: return datetime.fromisoformat(ts_str.replace("Z","+00:00")).astimezone(IST).strftime("%d-%m-%Y %H:%M:%S")
    except: return ts_str

def to_ms(ts_str):
    if not ts_str: return None
    try:
        dt = datetime.strptime(ts_str, "%a %b %d %H:%M:%S +0000 %Y")
        return int(dt.replace(tzinfo=timezone.utc).timestamp() * 1000)
    except: pass
    try: return int(datetime.fromisoformat(ts_str.replace("Z","+00:00")).timestamp() * 1000)
    except: return None

def to_ist_display(ts_str):
    if not ts_str: return "NA"
    try:
        dt = datetime.strptime(ts_str, "%a %b %d %H:%M:%S +0000 %Y")
        return dt.replace(tzinfo=timezone.utc).astimezone(IST).strftime("%d/%m/%Y %H:%M")
    except: pass
    try:
        return datetime.fromisoformat(ts_str.replace("Z","+00:00")).astimezone(IST).strftime("%d/%m/%Y %H:%M")
    except: return "NA"

def fmt_time(m):
    if m is None: return "—"
    if m >= 1440: return f"{round(m/1440)}d"
    if m >= 60:   return f"{m/60:.1f}h"
    return f"{m}m"

def mins(a, b):
    return round(abs(b - a) / 60000)

def frt_hours(ms1, ms2):
    if ms1 is None or ms2 is None: return None
    return round(abs(ms2 - ms1) / 3600000, 4)

# ──────────────── API (retry-hardened) ───────────────────

def api_get(path, params={}):
    """GET with rate-limit wait + automatic retries on timeouts,
    connection errors, 429 and 5xx. One flaky call no longer kills
    an insurer or a thread."""
    url = f"https://api.twitterapi.io{path}"
    attempts = 4
    for attempt in range(1, attempts + 1):
        time.sleep(WAIT)
        try:
            r = requests.get(url, headers={"X-API-Key": API_KEY},
                             params=params, timeout=30)
        except (requests.exceptions.Timeout,
                requests.exceptions.ConnectionError) as e:
            if attempt == attempts:
                raise
            wait = 10 * attempt
            print(f"    network issue ({type(e).__name__}) — retry {attempt}/{attempts-1} in {wait}s...")
            time.sleep(wait)
            continue
        if r.status_code == 429 or r.status_code >= 500:
            if attempt == attempts:
                r.raise_for_status()
            wait = 15 * attempt
            print(f"    HTTP {r.status_code} — retry {attempt}/{attempts-1} in {wait}s...")
            time.sleep(wait)
            continue
        r.raise_for_status()
        return r.json()

def parse_tweet(t):
    if not t: return None
    a = t.get("author") or {}
    handle = (a.get("userName") or a.get("screen_name") or "").lower()
    ts_str = t.get("createdAt") or t.get("created_at") or ""
    text   = t.get("text") or t.get("full_text") or ""
    ms     = to_ms(ts_str)
    return {"handle": handle, "ts_str": ts_str, "ts": ms, "text": text} if ms else None

def tweet_id_of(t):
    return str(t.get("id") or t.get("id_str") or t.get("tweetId") or "")

def is_root_tweet(t):
    if t.get("isReply") is True:
        return False
    if t.get("inReplyToId") or t.get("in_reply_to_status_id_str") or t.get("in_reply_to_status_id"):
        return False
    conv = t.get("conversationId") or t.get("conversation_id_str") or t.get("conversation_id")
    tid  = tweet_id_of(t)
    if conv and tid and str(conv) != tid:
        return False
    return True

def match_keywords(text):
    low = text.lower()
    return [kw for kw in PRIORITY_KEYWORDS if kw.lower() in low]

def is_media_author(t):
    a = t.get("author") or {}
    handle = (a.get("userName") or a.get("screen_name") or "").lower()
    if handle in EXCLUDE_AUTHORS:
        return True
    if SKIP_BUSINESS_VERIFIED and (a.get("verifiedType") or "").lower() in ("business", "government"):
        return True
    followers = a.get("followers") or a.get("followersCount") or a.get("followers_count") or 0
    if MAX_AUTHOR_FOLLOWERS and followers > MAX_AUTHOR_FOLLOWERS:
        return True
    return False

# ─────────────────────────────────────────────────────────
#  DISCOVERY SOURCE A — INSURER'S OWN REPLIES (the reliable one)
# ─────────────────────────────────────────────────────────
# Twitter's search index silently drops tweets from small accounts —
# which is exactly who complains. But the insurer's OWN timeline
# always shows their replies, and every reply carries the
# conversationId of the customer's root tweet. So: walk the
# insurer's timeline, collect conversation IDs from their replies,
# and batch-fetch those root tweets. This recovers every thread the
# insurer replied to, even when search cannot see the complaint.

def fetch_insurer_reply_conversations(handle, since_unix):
    """Walk @handle's own timeline (newest→oldest) and return the set
    of conversationIds of tweets they replied to, until we pass the
    window start."""
    conv_ids, cursor = set(), None
    for page in range(1, MAX_TIMELINE_PAGES + 1):
        params = {"userName": handle, "includeReplies": True}
        if cursor:
            params["cursor"] = cursor
        try:
            data = api_get("/twitter/user/last_tweets", params)
        except requests.exceptions.RequestException as e:
            print(f"    timeline @{handle} error: {e} — using what we have")
            break
        tweets = (data.get("tweets") or data.get("data") or [])
        # some responses nest under data.tweets
        if isinstance(tweets, dict):
            tweets = tweets.get("tweets") or []
        if not tweets:
            break
        oldest_ms = None
        n_replies = 0
        for t in tweets:
            ms = to_ms(t.get("createdAt") or t.get("created_at") or "")
            if ms is not None:
                oldest_ms = ms if oldest_ms is None else min(oldest_ms, ms)
            conv = t.get("conversationId") or t.get("conversation_id_str") or ""
            tid  = tweet_id_of(t)
            # a reply into someone else's conversation
            if conv and tid and str(conv) != tid:
                conv_ids.add(str(conv))
                n_replies += 1
        print(f"    @{handle} timeline page {page}: {len(tweets)} tweets, "
              f"{n_replies} replies (conversations so far: {len(conv_ids)})")
        # stop when the page is entirely older than the window start
        if oldest_ms is not None and oldest_ms < since_unix * 1000:
            break
        cursor = data.get("next_cursor")
        if not cursor or data.get("has_next_page") is False:
            break
    return conv_ids

def batch_fetch_tweets(ids):
    """Fetch tweets in batches via /twitter/tweets (tweet_ids)."""
    out = []
    ids = [i for i in ids if i]
    for i in range(0, len(ids), 20):
        chunk = ids[i:i+20]
        try:
            data = api_get("/twitter/tweets", {"tweet_ids": ",".join(chunk)})
        except requests.exceptions.RequestException as e:
            print(f"    batch fetch error: {e} — skipping {len(chunk)} ids")
            continue
        out.extend(data.get("tweets") or data.get("data") or [])
    return out

# ─────────────────────────────────────────────────────────
#  DISCOVERY SOURCE B — ADVANCED SEARCH (catches unreplied tweets
#  that ARE in the search index)
# ─────────────────────────────────────────────────────────

def fetch_search_for_handles(insurer_handles, since_unix, until_unix, cap):
    mention_part = "(" + " OR ".join(f"@{h}" for h in sorted(insurer_handles)) + ")"
    exclude_part = " ".join(f"-from:{h}" for h in sorted(insurer_handles))
    query = f"{mention_part} {exclude_part} since_time:{since_unix} until_time:{until_unix}"
    out, cursor, page = [], None, 0
    while len(out) < cap and page < MAX_SEARCH_PAGES:
        page += 1
        params = {"query": query, "queryType": "Latest"}
        if cursor:
            params["cursor"] = cursor
        try:
            data = api_get("/twitter/tweet/advanced_search", params)
        except requests.exceptions.RequestException as e:
            print(f"    search error: {e} — using what we have")
            break
        tweets = data.get("tweets") or data.get("data") or []
        if not tweets:
            break
        out.extend(tweets)
        print(f"    search page {page}: {len(tweets)} tweets (total {len(out)})")
        cursor = data.get("next_cursor")
        if not cursor or data.get("has_next_page") is False:
            break
    return out

# ─────────────────────────────────────────────────────────

def discover_tweets(insurer_handles):
    """Union of: (A) roots of threads the insurer replied to, taken
    from the insurer's own timeline, and (B) advanced search."""
    since_unix = date_to_unix(SEARCH_SINCE)
    until_unix = date_to_unix(SEARCH_UNTIL)

    print("  [Source A] Insurer's own replies → conversation roots:")
    conv_ids = set()
    for handle in sorted(insurer_handles):
        conv_ids |= fetch_insurer_reply_conversations(handle, since_unix)
    print(f"    {len(conv_ids)} unique conversations found from insurer replies")
    raw_roots = batch_fetch_tweets(sorted(conv_ids)) if conv_ids else []

    print("  [Source B] Advanced search:")
    raw_search = fetch_search_for_handles(insurer_handles, since_unix,
                                          until_unix, MAX_SEARCH_TWEETS * 3)

    root_ids   = {tweet_id_of(t) for t in raw_roots if tweet_id_of(t)}
    search_ids = {tweet_id_of(t) for t in raw_search if tweet_id_of(t)}
    print(f"  Sources: insurer-replies={len(root_ids)}, search={len(search_ids)}, "
          f"recovered-only-by-replies={len(root_ids - search_ids)}")

    found, seen = [], set()
    skipped_replies = skipped_media = skipped_window = 0

    for t, src in [(t, "insurer-replies") for t in raw_roots] + \
                  [(t, "search") for t in raw_search]:
        tid = tweet_id_of(t)
        if not tid or tid in seen:
            continue
        seen.add(tid)
        a = t.get("author") or {}
        author = (a.get("userName") or a.get("screen_name") or "user")
        if author.lower() in insurer_handles:
            continue
        if is_media_author(t):
            skipped_media += 1
            continue
        ms = to_ms(t.get("createdAt") or t.get("created_at") or "")
        if ms is None or not (since_unix * 1000 <= ms < until_unix * 1000):
            skipped_window += 1
            continue
        if ROOT_TWEETS_ONLY and not is_root_tweet(t):
            skipped_replies += 1
            continue
        found.append({
            "url": f"https://x.com/{author}/status/{tid}",
            "id": tid,
            "tweet": t,
            "source": src,
        })
        if len(found) >= MAX_SEARCH_TWEETS:
            break

    for item in found:
        text = (item["tweet"].get("text") or item["tweet"].get("full_text") or "")
        item["matched"] = match_keywords(text)
    found.sort(key=lambda x: 0 if x["matched"] else 1)
    n_pri = sum(1 for x in found if x["matched"])

    if skipped_media:
        print(f"  Skipped {skipped_media} media/sponsor tweets")
    if skipped_window:
        print(f"  Skipped {skipped_window} tweets outside the date window")
    if ROOT_TWEETS_ONLY and skipped_replies:
        print(f"  Skipped {skipped_replies} reply tweets (kept root tweets only)")
    print(f"  Priority (keyword) tweets: {n_pri} | Other tweets: {len(found) - n_pri}")
    return found

# ──────────────── THREADS — FULL CONTEXT ─────────────────

def fetch_thread(tid):
    all_tweets, cursor = [], None
    for _ in range(MAX_THREAD_PAGES):
        params = {"tweetId": tid}
        if cursor:
            params["cursor"] = cursor
        data = api_get("/twitter/tweet/thread_context", params)
        batch = data.get("replies") or data.get("tweets") or data.get("data") or []
        if not batch:
            break
        all_tweets.extend(batch)
        cursor = data.get("next_cursor")
        if not cursor or data.get("has_next_page") is False:
            break
    return all_tweets

# ──────────────── row builders ───────────────────────────

def build_scoring_row(insurer_name, url, posts):
    cust_ts, ins_ts, seen_c, seen_i = [], [], set(), set()
    for post in posts:
        ts_disp = to_ist_display(post["ts_str"])
        if post["role"] == "customer" and ts_disp not in seen_c:
            cust_ts.append(ts_disp); seen_c.add(ts_disp)
        elif post["role"] == "insurer" and ts_disp not in seen_i:
            ins_ts.append(ts_disp); seen_i.add(ts_disp)

    ts_cols = []
    for i in range(MAX_TIMESTAMPS):
        ts_cols.append(cust_ts[i] if i < len(cust_ts) else "NA")
        ts_cols.append(ins_ts[i]  if i < len(ins_ts)  else "NA")

    cust_ms = posts[0]["ts"] if posts else None
    ins_ms  = next((p["ts"] for p in posts if p["role"] == "insurer"), None)
    frt_val  = frt_hours(cust_ms, ins_ms) if (cust_ms and ins_ms) else None
    frt_disp = round(frt_val, 4) if frt_val is not None else "NA"

    frt_lt1 = frt_1to3 = frt_gt3 = "FALSE"
    points = ""
    if frt_val is not None:
        d = frt_val / 24
        if d < 1:      frt_lt1  = "TRUE"; points = 10
        elif d <= 3:   frt_1to3 = "TRUE"; points = 5
        else:          frt_gt3  = "TRUE"; points = -10

    score_note = "FRT < 1 day : 10 | 1 to 3 days : 5 | > 3 days : -10"
    return [insurer_name, url] + ts_cols + [frt_disp, frt_lt1, frt_1to3, frt_gt3, points, "", score_note]

def scoring_headers():
    row1 = ["Insurer", "Link"]
    for i in range(1, MAX_TIMESTAMPS + 1):
        suffix = {1:"st",2:"nd",3:"rd"}.get(i,"th")
        row1 += [f"{i}{suffix} Time Stamp", ""]
    row1 += ["FRT (First Response Time) (in Hrs)",
             "FRT less than 1 day (10)", "FRT >1 Day <3 days (5)",
             "FRT more than 3 days (-10)", "Points", "Marks", "Score Methodology"]
    row2 = ["", ""]
    for _ in range(MAX_TIMESTAMPS):
        row2 += ["Customer", "Insurer"]
    row2 += ["", "", "", "", "", "", ""]
    return row1, row2

URLS_HEADER   = ["Tweet URL", "Handle", "Tweet Time (IST)", "Found Via",
                 "Priority", "Matched Keywords", "Text (first 150 chars)"]
DETAIL_HEADER = ["Month", "Customer Tweet URL", "Insurer Replied?",
                 "First Response Time", "Avg Response Time",
                 "Sequence", "Role", "Handle",
                 "Timestamp IST", "Gap from prev (mins)", "Tweet text"]

# ──────────────── xlsx saving ────────────────────────────

def sheet_name(name):
    clean = re.sub(r'[\[\]:*?/\\]', '', name).strip()
    return clean[:31] if clean else "Sheet"

def write_workbook(path, per_insurer_rows, header_rows):
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in per_insurer_rows.items():
        ws = wb.create_sheet(sheet_name(name))
        n_header = len(header_rows)
        for hr in header_rows:
            ws.append(hr)
        for r in range(1, n_header + 1):
            for cell in ws[r]:
                cell.font = BOLD
                cell.fill = HEADER_FILL
        for row in rows:
            ws.append(row)
        ws.freeze_panes = ws.cell(row=n_header + 1, column=1)
    if not wb.sheetnames:
        wb.create_sheet("Empty")
    wb.save(path)

def save_all(urls_by, scoring_by, detail_by, json_rows):
    write_workbook(OUTPUT_URLS,   urls_by,    [URLS_HEADER])
    write_workbook(OUTPUT_MAIN,   scoring_by, list(scoring_headers()))
    write_workbook(OUTPUT_DETAIL, detail_by,  [DETAIL_HEADER])
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(json_rows, f, indent=2, ensure_ascii=False)

# ──────────────── per-insurer processing ─────────────────

def process_insurer(insurer, urls_by, scoring_by, detail_by, json_rows, stats):
    name    = insurer["name"]
    handles = {h.lower().replace("@","").strip() for h in insurer["handles"]}

    urls_by.setdefault(name, [])
    scoring_by.setdefault(name, [])
    detail_by.setdefault(name, [])

    print(f"\n{'─'*60}")
    print(f"INSURER: {name}  ({', '.join('@'+h for h in sorted(handles))})")
    print(f"{'─'*60}")

    items = discover_tweets(handles)
    if not items:
        print("  No tweets found for this insurer in the date window.")
        stats.append({"name": name, "found": 0, "processed": 0, "replied": 0, "avg_first": None})
        return

    for item in items:
        t = item["tweet"]
        a = t.get("author") or {}
        urls_by[name].append([
            item["url"],
            a.get("userName") or a.get("screen_name") or "",
            to_ist(t.get("createdAt") or t.get("created_at") or ""),
            item.get("source", ""),
            "Yes" if item.get("matched") else "No",
            ", ".join(item.get("matched") or []),
            (t.get("text") or t.get("full_text") or "").replace("\n"," ")[:150],
        ])

    print(f"  Found {len(items)} customer tweets. Fetching full threads...")

    success = replied = 0
    first_resp_times = []
    replier_threads = defaultdict(set)
    root_authors    = set()

    for i, item in enumerate(items):
        url, tid = item["url"], item["id"]
        star = "★ " if item.get("matched") else ""
        print(f"  [{i+1}/{len(items)}] {star}{tid}...", end=" ")
        try:
            posts = []
            root_p = parse_tweet(item["tweet"])
            root_handle = root_p["handle"] if root_p else ""
            root_authors.add(root_handle)
            if root_p:
                root_p["role"] = "customer"
                posts.append(root_p)

            thread = fetch_thread(tid)
            seen_ids = {tid}
            for t in thread:
                rid = tweet_id_of(t)
                if rid and rid in seen_ids:
                    continue
                if rid:
                    seen_ids.add(rid)
                p = parse_tweet(t)
                if p:
                    p["role"] = "insurer" if p["handle"] in handles else "customer"
                    posts.append(p)
                    if p["handle"] not in handles and p["handle"] != root_handle:
                        replier_threads[p["handle"]].add(tid)

            posts.sort(key=lambda x: x["ts"])
            if not posts:
                print("no data")
                continue

            cust    = posts[0]
            f_ins   = next((p for p in posts if p["role"] == "insurer"), None)
            f_resp  = mins(cust["ts"], f_ins["ts"]) if f_ins else None
            th_rep  = f_ins is not None

            ins_gaps = [mins(posts[j-1]["ts"], p["ts"])
                        for j, p in enumerate(posts) if p["role"] == "insurer" and j > 0]
            avg_th = round(sum(ins_gaps)/len(ins_gaps)) if ins_gaps else None
            mon    = datetime.fromtimestamp(cust["ts"]/1000, tz=IST).strftime("%b %Y")

            scoring_by[name].append(build_scoring_row(name, url, posts))

            for j, p in enumerate(posts):
                prev_ts = posts[j-1]["ts"] if j > 0 else None
                gap     = mins(prev_ts, p["ts"]) if prev_ts else ""
                detail_by[name].append([
                    mon, url, "Yes" if th_rep else "No",
                    fmt_time(f_resp) if j == 0 else "",
                    fmt_time(avg_th) if j == 0 else "",
                    j + 1, p["role"], p["handle"],
                    to_ist(p["ts_str"]), gap,
                    p["text"].replace("\n"," ").replace("\r","")[:200]
                ])

            for j, p in enumerate(posts):
                json_rows.append({
                    "insurer": name,
                    "postUrl": url, "postId": tid,
                    "replyId": f"{tid}_{j}", "replyUrl": url,
                    "replyText": p["text"], "timestamp": p["ts"],
                    "conversationId": tid,
                    "inReplyTo": None if j == 0 else tid,
                    "author": {"name": p["handle"], "screenName": p["handle"],
                               "followersCount":0,"favouritesCount":0,
                               "friendsCount":0,"description":""}
                })

            success += 1
            if th_rep:
                replied += 1
                if f_resp: first_resp_times.append(f_resp)
            print(f"✓ {len(posts)} turns | {'replied '+fmt_time(f_resp) if th_rep else 'no reply'}")

            if success % 5 == 0:
                save_all(urls_by, scoring_by, detail_by, json_rows)
                print(f"    💾 Checkpoint saved")

        except requests.exceptions.HTTPError as e:
            print(f"HTTP {e.response.status_code}: {e.response.text[:100]}")
        except Exception as e:
            print(f"Error: {e}")

    candidates = sorted(
        ((h, len(tids)) for h, tids in replier_threads.items()
         if len(tids) >= 2 and h not in root_authors),
        key=lambda x: -x[1])[:5]
    if candidates:
        print(f"\n  ⚠ Handles replying in multiple threads (NOT in your config):")
        for h, n in candidates:
            print(f"      @{h} — replied in {n} threads")
        print(f"    If any is the insurer's care/group handle, add it to")
        print(f"    '{name}' handles and re-run for correct reply detection.")

    avg_f = round(sum(first_resp_times)/len(first_resp_times)) if first_resp_times else None
    stats.append({"name": name, "found": len(items), "processed": success,
                  "replied": replied, "avg_first": avg_f})
    save_all(urls_by, scoring_by, detail_by, json_rows)
    print(f"  💾 {name} done — saved.")

# ──────────────── main ───────────────────────────────────

def main():
    print(f"\n{'='*60}")
    print(f"█ VERSION 12 █  Twitter Thread Fetcher — twitterapi.io")
    print(f"(insurer-replies discovery + search | retries on timeout)")
    print(f"If this banner does not say VERSION 12, you are running")
    print(f"an OLD file — check your filename and folder.")
    print(f"{'='*60}")
    print(f"Insurers in batch : {len(INSURERS)}")
    print(f"Date window       : {SEARCH_SINCE} → {SEARCH_UNTIL} (until exclusive)")
    print(f"Root tweets only  : {ROOT_TWEETS_ONLY}")
    print(f"{'='*60}")

    urls_by, scoring_by, detail_by, json_rows = {}, {}, {}, []
    stats = []

    for insurer in INSURERS:
        try:
            process_insurer(insurer, urls_by, scoring_by, detail_by, json_rows, stats)
        except Exception as e:
            print(f"  Fatal error for {insurer['name']}: {e} — moving to next insurer.")

    save_all(urls_by, scoring_by, detail_by, json_rows)

    print(f"\n{'='*60}")
    print(f"BATCH SUMMARY")
    print(f"{'='*60}")
    print(f"{'Insurer':30} {'Found':>6} {'Done':>6} {'Replied':>8} {'Avg 1st':>8}")
    for s in stats:
        rate = f"{round(s['replied']/s['processed']*100)}%" if s['processed'] else "—"
        print(f"{s['name'][:30]:30} {s['found']:>6} {s['processed']:>6} "
              f"{rate:>8} {fmt_time(s['avg_first']):>8}")

    import os
    print(f"\nFiles in: {os.getcwd()}")
    print(f"  {OUTPUT_URLS}     ← tweets found per insurer (with Found Via column)")
    print(f"  {OUTPUT_MAIN}     ← scoring workbook (1 sheet per insurer)")
    print(f"  {OUTPUT_DETAIL}   ← full thread detail (1 sheet per insurer)")
    print(f"  {OUTPUT_JSON}     ← upload to Phase 2 tool (has 'insurer' field)")
    print(f"{'='*60}\n")

if __name__ == "__main__":
    main()
